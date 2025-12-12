from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, session, send_from_directory
from flask_sqlalchemy import SQLAlchemy
from flask_socketio import SocketIO, emit, join_room, leave_room
from datetime import datetime, timedelta
from functools import wraps
from werkzeug.utils import secure_filename
import os
import uuid
import zipfile
import shutil
import random
from openpyxl import load_workbook

# Get the directory where this script is located
basedir = os.path.abspath(os.path.dirname(__file__))

app = Flask(__name__, template_folder=os.path.join(basedir, 'templates'), 
            static_folder=os.path.join(basedir, 'static'))
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'your-secret-key-change-in-production')
# Database configuration - use environment variable for production, sqlite for local
database_url = os.environ.get('DATABASE_URL', 'sqlite:///football_auction.db')
# Render uses postgres:// but SQLAlchemy needs postgresql://
if database_url.startswith('postgres://'):
    database_url = database_url.replace('postgres://', 'postgresql://', 1)
app.config['SQLALCHEMY_DATABASE_URI'] = database_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = 'static/uploads/team_logos'
app.config['PLAYER_PHOTO_FOLDER'] = 'static/uploads/player_photos'
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max file size (for bulk uploads)
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp', 'svg'}

db = SQLAlchemy(app)

# Initialize Socket.IO - use threading mode (compatible with Python 3.13)
# Eventlet doesn't support Python 3.13, so we use threading mode everywhere
socketio = SocketIO(app, cors_allowed_origins="*", async_mode='threading')

# Database Models
class Team(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    owner = db.Column(db.String(100), nullable=False)
    coowner_name = db.Column(db.String(100), nullable=True)
    batch = db.Column(db.String(50), nullable=False)
    price = db.Column(db.Float, default=0.0, nullable=False)
    number_of_members = db.Column(db.Integer, default=12, nullable=False)
    logo_filename = db.Column(db.String(255), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class AuctionSetting(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    auction_start_time = db.Column(db.DateTime, nullable=False)
    auction_date = db.Column(db.String(100), nullable=True)  # e.g., "Saturday, December 14, 2025"
    auction_place = db.Column(db.String(200), nullable=True)  # Auction venue/location
    is_active = db.Column(db.Boolean, default=True)

class Admin(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(50), unique=True, nullable=False)
    password = db.Column(db.String(100), nullable=False)

class TeamUser(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    team_id = db.Column(db.Integer, db.ForeignKey('team.id'), nullable=False)
    username = db.Column(db.String(50), unique=True, nullable=False)
    password = db.Column(db.String(100), nullable=False)
    team = db.relationship('Team', backref=db.backref('users', lazy=True))

class SlotManagement(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    total_slots = db.Column(db.Integer, default=12, nullable=False)
    total_teams = db.Column(db.Integer, default=12, nullable=False)  # Maximum teams allowed
    filled_slots = db.Column(db.Integer, default=0, nullable=False)
    remaining_slots = db.Column(db.Integer, default=12, nullable=False)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

class Player(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    batch = db.Column(db.String(50), nullable=False)
    position = db.Column(db.String(50), nullable=False)
    base_price = db.Column(db.Float, default=0.0, nullable=False)
    photo_filename = db.Column(db.String(255), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

# Many-to-many relationship table for Auction and Player
auction_players = db.Table('auction_players',
    db.Column('auction_id', db.Integer, db.ForeignKey('auction.id'), primary_key=True),
    db.Column('player_id', db.Integer, db.ForeignKey('player.id'), primary_key=True)
)

class Auction(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False)
    min_bid = db.Column(db.Float, default=0.0, nullable=False)
    max_bid = db.Column(db.Float, nullable=True)
    sponsor = db.Column(db.String(200), nullable=True)
    status = db.Column(db.String(20), default='draft', nullable=False)  # draft, live, closed
    auction_setting_id = db.Column(db.Integer, db.ForeignKey('auction_setting.id'), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    # Live auction state fields
    is_live = db.Column(db.Boolean, default=False, nullable=False)
    current_player_id = db.Column(db.Integer, db.ForeignKey('player.id'), nullable=True)
    highest_bid = db.Column(db.Float, default=0.0, nullable=False)
    highest_bid_team_id = db.Column(db.Integer, db.ForeignKey('team.id'), nullable=True)
    
    # Relationships
    auction_setting = db.relationship('AuctionSetting', backref=db.backref('auctions', lazy=True))
    players = db.relationship('Player', secondary=auction_players, lazy='subquery',
                           backref=db.backref('auctions', lazy=True))
    current_player = db.relationship('Player', foreign_keys=[current_player_id], lazy=True)
    highest_bid_team = db.relationship('Team', foreign_keys=[highest_bid_team_id], lazy=True)

class Bid(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    auction_id = db.Column(db.Integer, db.ForeignKey('auction.id'), nullable=False)
    player_id = db.Column(db.Integer, db.ForeignKey('player.id'), nullable=False)
    team_id = db.Column(db.Integer, db.ForeignKey('team.id'), nullable=False)
    bid_amount = db.Column(db.Float, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    # Relationships
    auction = db.relationship('Auction', backref=db.backref('bids', lazy=True))
    player = db.relationship('Player', backref=db.backref('bids', lazy=True))
    team = db.relationship('Team', backref=db.backref('bids', lazy=True))

# Authentication decorators
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'admin_logged_in' not in session:
            flash('Please login as admin first', 'error')
            return redirect(url_for('admin_login'))
        return f(*args, **kwargs)
    return decorated_function

def team_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'team_logged_in' not in session:
            flash('Please login as team first', 'error')
            return redirect(url_for('team_login'))
        return f(*args, **kwargs)
    return decorated_function

# Helper function to check allowed file extensions
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# Helper function to save uploaded file
def save_team_logo(file):
    if file and allowed_file(file.filename):
        # Generate unique filename
        filename = secure_filename(file.filename)
        unique_filename = f"{uuid.uuid4().hex}_{filename}"
        
        # Create upload directory if it doesn't exist
        upload_folder = app.config['UPLOAD_FOLDER']
        os.makedirs(upload_folder, exist_ok=True)
        
        # Save file
        file_path = os.path.join(upload_folder, unique_filename)
        file.save(file_path)
        return unique_filename
    return None

# Helper function to save player photo
def save_player_photo(file):
    if file and allowed_file(file.filename):
        # Generate unique filename
        filename = secure_filename(file.filename)
        unique_filename = f"{uuid.uuid4().hex}_{filename}"
        
        # Create upload directory if it doesn't exist
        upload_folder = app.config['PLAYER_PHOTO_FOLDER']
        os.makedirs(upload_folder, exist_ok=True)
        
        # Save file
        file_path = os.path.join(upload_folder, unique_filename)
        file.save(file_path)
        return unique_filename
    return None

# Football positions list
FOOTBALL_POSITIONS = [
    'Goalkeeper (GK)',
    'Right Back (RB)',
    'Left Back (LB)',
    'Center Back (CB)',
    'Defensive Midfielder (CDM)',
    'Right Midfielder (RM)',
    'Left Midfielder (LM)',
    'Central Midfielder (CM)',
    'Attacking Midfielder (CAM)',
    'Right Winger (RW)',
    'Left Winger (LW)',
    'Striker (ST)',
    'Center Forward (CF)',
    'Second Striker (SS)'
]

# Routes
@app.route('/')
def home():
    teams = Team.query.all()
    team_count = len(teams)
    players = Player.query.all()
    player_count = len(players)
    
    # Check for live auction
    live_auction = Auction.query.filter_by(is_live=True, status='live').first()
    
    # Get slot information
    slot_info = SlotManagement.query.first()
    if not slot_info:
        slot_info = SlotManagement(total_slots=12, total_teams=12, filled_slots=0, remaining_slots=12)
        db.session.add(slot_info)
        db.session.commit()
    
    # Update filled slots based on actual teams
    slot_info.filled_slots = team_count
    slot_info.remaining_slots = slot_info.total_slots - team_count
    db.session.commit()
    
    # Get auction settings from database
    auction_setting = AuctionSetting.query.first()
    if not auction_setting:
        # Default: Saturday December 14, 2025 at 10:00 AM
        countdown_time = datetime(2025, 12, 14, 10, 0, 0)
        auction_setting = AuctionSetting(
            auction_start_time=countdown_time,
            auction_date="Saturday, December 14, 2025",
            auction_place="Main Auditorium"
        )
        db.session.add(auction_setting)
        db.session.commit()
    
    countdown_time = auction_setting.auction_start_time
    
    return render_template('index.html', teams=teams, team_count=team_count, 
                         countdown_time=countdown_time, slot_info=slot_info,
                         players=players, player_count=player_count,
                         auction_setting=auction_setting, live_auction=live_auction)

@app.route('/api/countdown')
def get_countdown():
    auction_setting = AuctionSetting.query.first()
    if auction_setting:
        countdown_time = auction_setting.auction_start_time
        now = datetime.utcnow()
        if countdown_time > now:
            time_left = countdown_time - now
            return jsonify({
                'days': time_left.days,
                'hours': time_left.seconds // 3600,
                'minutes': (time_left.seconds % 3600) // 60,
                'seconds': time_left.seconds % 60,
                'total_seconds': int(time_left.total_seconds())
            })
    return jsonify({'error': 'No auction time set'}), 404

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        admin = Admin.query.filter_by(username=username).first()
        if admin and admin.password == password:
            session['admin_logged_in'] = True
            session['admin_username'] = username
            flash('Admin login successful!', 'success')
            return redirect(url_for('admin_dashboard'))
        else:
            flash('Invalid credentials', 'error')
    
    return render_template('admin_login.html')

@app.route('/admin/logout')
@admin_required
def admin_logout():
    session.pop('admin_logged_in', None)
    session.pop('admin_username', None)
    flash('Admin logged out successfully', 'success')
    return redirect(url_for('home'))

@app.route('/team/login', methods=['GET', 'POST'])
def team_login():
    if request.method == 'POST':
        team_name = request.form.get('team_name')
        password = request.form.get('password')
        
        # Find team by name
        team = Team.query.filter_by(name=team_name).first()
        if team:
            # Check if team has a user account
            team_user = TeamUser.query.filter_by(team_id=team.id).first()
            if team_user and team_user.password == password:
                session['team_logged_in'] = True
                session['team_id'] = team.id
                session['team_name'] = team.name
                flash('Team login successful!', 'success')
                return redirect(url_for('team_dashboard', team_id=team.id))
            elif not team_user:
                # Create default password for team if no user exists
                # Default password is team name in lowercase + "123"
                default_password = team_name.lower().replace(' ', '') + '123'
                if password == default_password:
                    # Create team user
                    new_team_user = TeamUser(team_id=team.id, username=team_name.lower().replace(' ', ''), password=default_password)
                    db.session.add(new_team_user)
                    db.session.commit()
                    session['team_logged_in'] = True
                    session['team_id'] = team.id
                    session['team_name'] = team.name
                    flash('Team login successful!', 'success')
                    return redirect(url_for('team_dashboard', team_id=team.id))
        
        flash('Invalid team name or password', 'error')
    
    return render_template('team_login.html')

@app.route('/admin/dashboard')
@admin_required
def admin_dashboard():
    teams = Team.query.all()
    players = Player.query.all()
    auctions = Auction.query.order_by(Auction.created_at.desc()).all()
    auction_setting = AuctionSetting.query.first()
    slot_info = SlotManagement.query.first()
    if not slot_info:
        slot_info = SlotManagement(total_slots=12, total_teams=12, filled_slots=len(teams), remaining_slots=12-len(teams))
        db.session.add(slot_info)
        db.session.commit()
    else:
        slot_info.filled_slots = len(teams)
        slot_info.remaining_slots = slot_info.total_slots - len(teams)
        db.session.commit()
    return render_template('admin_dashboard.html', teams=teams, players=players, auctions=auctions, auction_setting=auction_setting, slot_info=slot_info)

@app.route('/admin/settings', methods=['GET', 'POST'])
@admin_required
def admin_settings():
    slot_info = SlotManagement.query.first()
    auction_setting = AuctionSetting.query.first()
    auctions = Auction.query.all()
    
    if not slot_info:
        slot_info = SlotManagement(total_slots=12, total_teams=12, filled_slots=0, remaining_slots=12)
        db.session.add(slot_info)
        db.session.commit()
    
    if not auction_setting:
        default_time = datetime(2025, 12, 14, 10, 0, 0)
        auction_setting = AuctionSetting(
            auction_start_time=default_time,
            auction_date="Saturday, December 14, 2025",
            auction_place="Main Auditorium"
        )
        db.session.add(auction_setting)
        db.session.commit()
    
    if request.method == 'POST':
        # Update slot management
        total_slots = request.form.get('total_slots')
        total_teams = request.form.get('total_teams')
        
        try:
            slot_info.total_slots = int(total_slots) if total_slots else 12
            slot_info.total_teams = int(total_teams) if total_teams else 12
            # Recalculate remaining slots
            teams = Team.query.all()
            slot_info.filled_slots = len(teams)
            slot_info.remaining_slots = slot_info.total_slots - slot_info.filled_slots
            db.session.commit()
        except ValueError:
            flash('Invalid slot or team numbers', 'error')
        
        # Update auction settings - check if specific auction is selected
        selected_auction_id = request.form.get('selected_auction_id')
        auction_date = request.form.get('auction_date')
        auction_time = request.form.get('auction_time')
        auction_place = request.form.get('auction_place')
        
        # If an auction is selected, update that auction's setting
        if selected_auction_id:
            selected_auction = Auction.query.get(int(selected_auction_id))
            if selected_auction:
                if not selected_auction.auction_setting_id:
                    # Create new auction setting for this auction
                    if auction_date and auction_time:
                        try:
                            date_time_str = f"{auction_date} {auction_time}"
                            countdown_time = datetime.strptime(date_time_str, "%Y-%m-%d %H:%M")
                            new_setting = AuctionSetting(
                                auction_start_time=countdown_time,
                                auction_date=auction_date,
                                auction_place=auction_place or "Main Auditorium"
                            )
                            db.session.add(new_setting)
                            db.session.flush()
                            selected_auction.auction_setting_id = new_setting.id
                            db.session.commit()
                            flash('Auction settings updated successfully!', 'success')
                        except ValueError:
                            flash('Invalid date or time format', 'error')
                    else:
                        flash('Please provide both date and time', 'error')
                else:
                    # Update existing setting
                    setting = AuctionSetting.query.get(selected_auction.auction_setting_id)
                    if setting and auction_date and auction_time:
                        try:
                            date_time_str = f"{auction_date} {auction_time}"
                            countdown_time = datetime.strptime(date_time_str, "%Y-%m-%d %H:%M")
                            setting.auction_start_time = countdown_time
                            setting.auction_date = auction_date
                            setting.auction_place = auction_place or "Main Auditorium"
                            db.session.commit()
                            flash('Auction settings updated successfully!', 'success')
                        except ValueError:
                            flash('Invalid date or time format', 'error')
                    else:
                        flash('Please provide both date and time', 'error')
        else:
            # Update default auction settings
            if auction_date and auction_time:
                try:
                    # Parse date and time
                    date_time_str = f"{auction_date} {auction_time}"
                    countdown_time = datetime.strptime(date_time_str, "%Y-%m-%d %H:%M")
                    
                    auction_setting.auction_start_time = countdown_time
                    auction_setting.auction_date = auction_date
                    auction_setting.auction_place = auction_place or "Main Auditorium"
                    db.session.commit()
                    
                    flash('Settings updated successfully!', 'success')
                except ValueError:
                    flash('Invalid date or time format', 'error')
            else:
                flash('Please provide both date and time', 'error')
        
        return redirect(url_for('admin_settings'))
    
    # Format datetime for form input
    auction_datetime = auction_setting.auction_start_time
    auction_date_str = auction_datetime.strftime('%Y-%m-%d')
    auction_time_str = auction_datetime.strftime('%H:%M')
    
    return render_template('admin_settings.html', 
                         slot_info=slot_info, 
                         auction_setting=auction_setting,
                         auctions=auctions,
                         auction_date_str=auction_date_str,
                         auction_time_str=auction_time_str)

@app.route('/team/dashboard/<int:team_id>')
@team_required
def team_dashboard(team_id):
    # Verify team belongs to logged in user
    if session.get('team_id') != team_id:
        flash('Unauthorized access', 'error')
        return redirect(url_for('team_login'))
    team = Team.query.get_or_404(team_id)
    return render_template('team_dashboard.html', team=team)

@app.route('/team/logout')
@team_required
def team_logout():
    session.pop('team_logged_in', None)
    session.pop('team_id', None)
    session.pop('team_name', None)
    flash('Team logged out successfully', 'success')
    return redirect(url_for('home'))

@app.route('/admin/team/add', methods=['GET', 'POST'])
@admin_required
def add_team():
    if request.method == 'POST':
        name = request.form.get('name')
        owner = request.form.get('owner')
        coowner_name = request.form.get('coowner_name', '')
        batch = request.form.get('batch')
        price = request.form.get('price', 0.0)
        number_of_members = request.form.get('number_of_members', 12)
        
        # Handle logo upload
        logo_filename = None
        if 'logo' in request.files:
            logo_file = request.files['logo']
            if logo_file and logo_file.filename:
                logo_filename = save_team_logo(logo_file)
                if not logo_filename:
                    flash('Invalid logo file. Allowed formats: PNG, JPG, JPEG, GIF, WEBP, SVG', 'error')
        
        # Validate required fields
        if not name or not owner or not batch:
            flash('Team name, owner name, and batch are required', 'error')
            return redirect(url_for('add_team'))
        
        # Check if team name already exists
        existing_team = Team.query.filter_by(name=name).first()
        if existing_team:
            flash('Team name already exists', 'error')
            return redirect(url_for('add_team'))
        
        try:
            price = float(price) if price else 0.0
        except ValueError:
            price = 0.0
        
        try:
            number_of_members = int(number_of_members) if number_of_members else 12
        except ValueError:
            number_of_members = 12
        
        # Create new team
        new_team = Team(
            name=name,
            owner=owner,
            coowner_name=coowner_name,
            batch=batch,
            price=price,
            number_of_members=number_of_members,
            logo_filename=logo_filename
        )
        db.session.add(new_team)
        db.session.commit()
        
        # Update slot management
        slot_info = SlotManagement.query.first()
        if slot_info:
            slot_info.filled_slots = Team.query.count()
            slot_info.remaining_slots = slot_info.total_slots - slot_info.filled_slots
            db.session.commit()
        
        flash('Team added successfully!', 'success')
        return redirect(url_for('admin_dashboard'))
    
    return render_template('add_team.html')

@app.route('/admin/team/edit/<int:team_id>', methods=['GET', 'POST'])
@admin_required
def edit_team(team_id):
    team = Team.query.get_or_404(team_id)
    
    if request.method == 'POST':
        name = request.form.get('name')
        owner = request.form.get('owner')
        coowner_name = request.form.get('coowner_name', '')
        batch = request.form.get('batch')
        price = request.form.get('price', 0.0)
        number_of_members = request.form.get('number_of_members', 12)
        delete_logo = request.form.get('delete_logo') == 'true'
        
        # Handle logo upload
        if 'logo' in request.files:
            logo_file = request.files['logo']
            if logo_file and logo_file.filename:
                # Delete old logo if exists
                if team.logo_filename:
                    old_logo_path = os.path.join(app.config['UPLOAD_FOLDER'], team.logo_filename)
                    if os.path.exists(old_logo_path):
                        os.remove(old_logo_path)
                
                # Save new logo
                logo_filename = save_team_logo(logo_file)
                if logo_filename:
                    team.logo_filename = logo_filename
                elif not logo_filename:
                    flash('Invalid logo file. Allowed formats: PNG, JPG, JPEG, GIF, WEBP, SVG', 'error')
        
        # Handle logo deletion
        if delete_logo and team.logo_filename:
            old_logo_path = os.path.join(app.config['UPLOAD_FOLDER'], team.logo_filename)
            if os.path.exists(old_logo_path):
                os.remove(old_logo_path)
            team.logo_filename = None
        
        # Validate required fields
        if not name or not owner or not batch:
            flash('Team name, owner name, and batch are required', 'error')
            return redirect(url_for('edit_team', team_id=team_id))
        
        # Check if team name already exists (excluding current team)
        existing_team = Team.query.filter_by(name=name).first()
        if existing_team and existing_team.id != team_id:
            flash('Team name already exists', 'error')
            return redirect(url_for('edit_team', team_id=team_id))
        
        try:
            price = float(price) if price else 0.0
        except ValueError:
            price = 0.0
        
        try:
            number_of_members = int(number_of_members) if number_of_members else 12
        except ValueError:
            number_of_members = 12
        
        # Update team
        team.name = name
        team.owner = owner
        team.coowner_name = coowner_name
        team.batch = batch
        team.price = price
        team.number_of_members = number_of_members
        db.session.commit()
        
        flash('Team updated successfully!', 'success')
        return redirect(url_for('admin_dashboard'))
    
    return render_template('edit_team.html', team=team)

@app.route('/admin/team/delete/<int:team_id>', methods=['POST'])
@admin_required
def delete_team(team_id):
    team = Team.query.get_or_404(team_id)
    
    # Delete team logo if exists
    if team.logo_filename:
        logo_path = os.path.join(app.config['UPLOAD_FOLDER'], team.logo_filename)
        if os.path.exists(logo_path):
            try:
                os.remove(logo_path)
            except:
                pass
    
    # Delete associated team users
    TeamUser.query.filter_by(team_id=team_id).delete()
    
    # Delete team
    db.session.delete(team)
    db.session.commit()
    
    # Update slot management
    slot_info = SlotManagement.query.first()
    if slot_info:
        slot_info.filled_slots = Team.query.count()
        slot_info.remaining_slots = slot_info.total_slots - slot_info.filled_slots
        db.session.commit()
    
    flash('Team deleted successfully!', 'success')
    return redirect(url_for('admin_dashboard'))

@app.route('/api/teams')
def get_teams():
    teams = Team.query.all()
    return jsonify([{
        'id': team.id,
        'name': team.name,
        'owner': team.owner,
        'batch': team.batch
    } for team in teams])

# Player Management Routes
@app.route('/admin/player/add', methods=['GET', 'POST'])
@admin_required
def add_player():
    if request.method == 'POST':
        name = request.form.get('name')
        batch = request.form.get('batch')
        position = request.form.get('position')
        base_price = request.form.get('base_price', 0.0)
        
        # Handle photo upload
        photo_filename = None
        if 'photo' in request.files:
            photo_file = request.files['photo']
            if photo_file and photo_file.filename:
                photo_filename = save_player_photo(photo_file)
                if not photo_filename:
                    flash('Invalid photo file. Allowed formats: PNG, JPG, JPEG, GIF, WEBP, SVG', 'error')
        
        # Validate required fields
        if not name or not batch or not position:
            flash('Player name, batch, and position are required', 'error')
            return redirect(url_for('add_player'))
        
        try:
            base_price = float(base_price) if base_price else 0.0
        except ValueError:
            base_price = 0.0
        
        # Create new player
        new_player = Player(
            name=name,
            batch=batch,
            position=position,
            base_price=base_price,
            photo_filename=photo_filename
        )
        db.session.add(new_player)
        db.session.commit()
        
        flash('Player added successfully!', 'success')
        return redirect(url_for('admin_dashboard'))
    
    return render_template('add_player.html', positions=FOOTBALL_POSITIONS)

@app.route('/admin/player/edit/<int:player_id>', methods=['GET', 'POST'])
@admin_required
def edit_player(player_id):
    player = Player.query.get_or_404(player_id)
    
    if request.method == 'POST':
        name = request.form.get('name')
        batch = request.form.get('batch')
        position = request.form.get('position')
        base_price = request.form.get('base_price', 0.0)
        delete_photo = request.form.get('delete_photo') == 'true'
        
        # Handle photo upload
        if 'photo' in request.files:
            photo_file = request.files['photo']
            if photo_file and photo_file.filename:
                # Delete old photo if exists
                if player.photo_filename:
                    old_photo_path = os.path.join(app.config['PLAYER_PHOTO_FOLDER'], player.photo_filename)
                    if os.path.exists(old_photo_path):
                        os.remove(old_photo_path)
                
                # Save new photo
                photo_filename = save_player_photo(photo_file)
                if photo_filename:
                    player.photo_filename = photo_filename
                elif not photo_filename:
                    flash('Invalid photo file. Allowed formats: PNG, JPG, JPEG, GIF, WEBP, SVG', 'error')
        
        # Handle photo deletion
        if delete_photo and player.photo_filename:
            old_photo_path = os.path.join(app.config['PLAYER_PHOTO_FOLDER'], player.photo_filename)
            if os.path.exists(old_photo_path):
                os.remove(old_photo_path)
            player.photo_filename = None
        
        # Validate required fields
        if not name or not batch or not position:
            flash('Player name, batch, and position are required', 'error')
            return redirect(url_for('edit_player', player_id=player_id))
        
        try:
            base_price = float(base_price) if base_price else 0.0
        except ValueError:
            base_price = 0.0
        
        # Update player
        player.name = name
        player.batch = batch
        player.position = position
        player.base_price = base_price
        db.session.commit()
        
        flash('Player updated successfully!', 'success')
        return redirect(url_for('admin_dashboard'))
    
    return render_template('edit_player.html', player=player, positions=FOOTBALL_POSITIONS)

@app.route('/admin/player/delete/<int:player_id>', methods=['POST'])
@admin_required
def delete_player(player_id):
    player = Player.query.get_or_404(player_id)
    
    # Delete player photo if exists
    if player.photo_filename:
        photo_path = os.path.join(app.config['PLAYER_PHOTO_FOLDER'], player.photo_filename)
        if os.path.exists(photo_path):
            try:
                os.remove(photo_path)
            except:
                pass
    
    # Delete player
    db.session.delete(player)
    db.session.commit()
    
    flash('Player deleted successfully!', 'success')
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/player/bulk-upload', methods=['GET', 'POST'])
@admin_required
def bulk_upload_players():
    if request.method == 'POST':
        excel_file = request.files.get('excel_file')
        photos_folder = request.files.getlist('photos_folder')
        photos_zip = request.files.get('photos_zip')
        
        if not excel_file or excel_file.filename == '':
            flash('Please upload an Excel file', 'error')
            return redirect(url_for('bulk_upload_players'))
        
        # Create temporary directory for photos
        temp_photos_dir = os.path.join('static', 'temp_photos', uuid.uuid4().hex)
        os.makedirs(temp_photos_dir, exist_ok=True)
        
        try:
            # Handle zip file upload
            if photos_zip and photos_zip.filename:
                if photos_zip.filename.endswith('.zip'):
                    zip_path = os.path.join(temp_photos_dir, 'photos.zip')
                    photos_zip.save(zip_path)
                    
                    # Extract zip file
                    with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                        zip_ref.extractall(temp_photos_dir)
                    os.remove(zip_path)
            
            # Handle multiple file uploads
            if photos_folder:
                for photo_file in photos_folder:
                    if photo_file and photo_file.filename:
                        filename = secure_filename(photo_file.filename)
                        photo_path = os.path.join(temp_photos_dir, filename)
                        photo_file.save(photo_path)
            
            # Read Excel file
            try:
                workbook = load_workbook(excel_file, data_only=True)
                sheet = workbook.active
            except Exception as e:
                flash(f'Error reading Excel file: {str(e)}', 'error')
                shutil.rmtree(temp_photos_dir, ignore_errors=True)
                return redirect(url_for('bulk_upload_players'))
            
            # Get header row (first row)
            headers = []
            for cell in sheet[1]:
                headers.append(str(cell.value).lower().strip() if cell.value else '')
            
            # Validate required columns
            required_columns = ['player_name', 'batch', 'position']
            missing_columns = [col for col in required_columns if col.lower() not in headers]
            if missing_columns:
                flash(f'Missing required columns in Excel: {", ".join(missing_columns)}', 'error')
                shutil.rmtree(temp_photos_dir, ignore_errors=True)
                return redirect(url_for('bulk_upload_players'))
            
            # Get column indices
            player_name_idx = headers.index('player_name') if 'player_name' in headers else -1
            batch_idx = headers.index('batch') if 'batch' in headers else -1
            position_idx = headers.index('position') if 'position' in headers else -1
            base_price_idx = headers.index('base_price') if 'base_price' in headers else -1
            photo_name_idx = headers.index('photo_name') if 'photo_name' in headers else -1
            
            # Process each row
            success_count = 0
            error_count = 0
            errors = []
            
            for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
                try:
                    # Get cell values
                    player_name = str(row[player_name_idx].value).strip() if player_name_idx >= 0 and row[player_name_idx].value else ''
                    batch = str(row[batch_idx].value).strip() if batch_idx >= 0 and row[batch_idx].value else ''
                    position = str(row[position_idx].value).strip() if position_idx >= 0 and row[position_idx].value else ''
                    base_price = row[base_price_idx].value if base_price_idx >= 0 and row[base_price_idx].value else 0.0
                    photo_name = str(row[photo_name_idx].value).strip() if photo_name_idx >= 0 and row[photo_name_idx].value else None
                    
                    # Validate required fields
                    if not player_name or not batch or not position:
                        error_count += 1
                        errors.append(f'Row {index}: Missing required fields')
                        continue
                    
                    # Convert base_price to float
                    try:
                        base_price = float(base_price) if base_price else 0.0
                    except (ValueError, TypeError):
                        base_price = 0.0
                    
                    # Handle photo upload
                    photo_filename = None
                    if photo_name and photo_name.strip():
                        photo_name_clean = photo_name.strip()
                        # Look for photo in temp directory (case-insensitive matching)
                        photo_path = None
                        
                        # First try exact match (case-insensitive)
                        for root, dirs, files in os.walk(temp_photos_dir):
                            for file in files:
                                if file.lower() == photo_name_clean.lower():
                                    photo_path = os.path.join(root, file)
                                    break
                            if photo_path:
                                break
                        
                        # If not found, try partial match (filename without extension)
                        if not photo_path:
                            photo_name_base = os.path.splitext(photo_name_clean)[0].lower()
                            for root, dirs, files in os.walk(temp_photos_dir):
                                for file in files:
                                    file_base = os.path.splitext(file)[0].lower()
                                    if file_base == photo_name_base:
                                        photo_path = os.path.join(root, file)
                                        break
                                if photo_path:
                                    break
                        
                        # If still not found, try contains match
                        if not photo_path:
                            for root, dirs, files in os.walk(temp_photos_dir):
                                for file in files:
                                    if photo_name_base in file.lower() or file.lower() in photo_name_base:
                                        photo_path = os.path.join(root, file)
                                        break
                                if photo_path:
                                    break
                        
                        if photo_path and os.path.exists(photo_path):
                            # Validate it's an image file
                            if allowed_file(photo_path):
                                # Save photo to permanent location
                                file_ext = os.path.splitext(photo_path)[1]
                                unique_filename = f"{uuid.uuid4().hex}{file_ext}"
                                dest_path = os.path.join(app.config['PLAYER_PHOTO_FOLDER'], unique_filename)
                                os.makedirs(app.config['PLAYER_PHOTO_FOLDER'], exist_ok=True)
                                shutil.copy2(photo_path, dest_path)
                                photo_filename = unique_filename
                            else:
                                errors.append(f'Row {index}: Invalid photo format for {photo_name}')
                        else:
                            # Photo not found, but continue without photo
                            pass
                    
                    # Create player
                    new_player = Player(
                        name=player_name,
                        batch=batch,
                        position=position,
                        base_price=base_price,
                        photo_filename=photo_filename
                    )
                    db.session.add(new_player)
                    success_count += 1
                    
                except Exception as e:
                    error_count += 1
                    errors.append(f'Row {index}: {str(e)}')
                    continue
            
            # Commit all players
            db.session.commit()
            
            # Clean up temp directory
            shutil.rmtree(temp_photos_dir, ignore_errors=True)
            
            # Show results
            if success_count > 0:
                flash(f'Successfully uploaded {success_count} player(s)!', 'success')
            if error_count > 0:
                error_msg = f'Failed to upload {error_count} player(s).'
                if errors:
                    error_msg += ' Errors: ' + '; '.join(errors[:5])  # Show first 5 errors
                    if len(errors) > 5:
                        error_msg += f' ... and {len(errors) - 5} more'
                flash(error_msg, 'error')
            
            return redirect(url_for('admin_dashboard'))
            
        except Exception as e:
            shutil.rmtree(temp_photos_dir, ignore_errors=True)
            flash(f'Error processing bulk upload: {str(e)}', 'error')
            return redirect(url_for('bulk_upload_players'))
    
    return render_template('bulk_upload_players.html', positions=FOOTBALL_POSITIONS)

# Auction Management Routes
@app.route('/admin/auction/add', methods=['GET', 'POST'])
@admin_required
def add_auction():
    if request.method == 'POST':
        name = request.form.get('name')
        min_bid = request.form.get('min_bid', 0.0)
        max_bid = request.form.get('max_bid')
        sponsor = request.form.get('sponsor', '')
        player_ids = request.form.getlist('players')  # Get list of selected player IDs
        auction_setting_id = request.form.get('auction_setting_id')
        
        # Validate required fields
        if not name:
            flash('Auction name is required', 'error')
            return redirect(url_for('add_auction'))
        
        try:
            min_bid = float(min_bid) if min_bid else 0.0
        except ValueError:
            min_bid = 0.0
        
        try:
            max_bid = float(max_bid) if max_bid else None
        except ValueError:
            max_bid = None
        
        # Create new auction
        new_auction = Auction(
            name=name,
            min_bid=min_bid,
            max_bid=max_bid,
            sponsor=sponsor,
            status='draft',
            auction_setting_id=int(auction_setting_id) if auction_setting_id else None
        )
        
        # Add selected players
        if player_ids:
            players = Player.query.filter(Player.id.in_([int(pid) for pid in player_ids])).all()
            new_auction.players = players
        
        db.session.add(new_auction)
        db.session.commit()
        
        flash('Auction created successfully!', 'success')
        return redirect(url_for('admin_dashboard'))
    
    # Get all players and auction settings for the form
    players = Player.query.all()
    auction_settings = AuctionSetting.query.all()
    return render_template('add_auction.html', players=players, auction_settings=auction_settings)

@app.route('/admin/auction/edit/<int:auction_id>', methods=['GET', 'POST'])
@admin_required
def edit_auction(auction_id):
    auction = Auction.query.get_or_404(auction_id)
    
    if request.method == 'POST':
        name = request.form.get('name')
        min_bid = request.form.get('min_bid', 0.0)
        max_bid = request.form.get('max_bid')
        sponsor = request.form.get('sponsor', '')
        player_ids = request.form.getlist('players')
        auction_setting_id = request.form.get('auction_setting_id')
        
        if not name:
            flash('Auction name is required', 'error')
            return redirect(url_for('edit_auction', auction_id=auction_id))
        
        try:
            min_bid = float(min_bid) if min_bid else 0.0
        except ValueError:
            min_bid = 0.0
        
        try:
            max_bid = float(max_bid) if max_bid else None
        except ValueError:
            max_bid = None
        
        # Update auction
        auction.name = name
        auction.min_bid = min_bid
        auction.max_bid = max_bid
        auction.sponsor = sponsor
        auction.auction_setting_id = int(auction_setting_id) if auction_setting_id else None
        
        # Update players
        if player_ids:
            players = Player.query.filter(Player.id.in_([int(pid) for pid in player_ids])).all()
            auction.players = players
        else:
            auction.players = []
        
        db.session.commit()
        
        flash('Auction updated successfully!', 'success')
        return redirect(url_for('admin_dashboard'))
    
    players = Player.query.all()
    auction_settings = AuctionSetting.query.all()
    selected_player_ids = [p.id for p in auction.players]
    return render_template('edit_auction.html', auction=auction, players=players, 
                         auction_settings=auction_settings, selected_player_ids=selected_player_ids)

@app.route('/admin/auction/delete/<int:auction_id>', methods=['POST'])
@admin_required
def delete_auction(auction_id):
    auction = Auction.query.get_or_404(auction_id)
    db.session.delete(auction)
    db.session.commit()
    flash('Auction deleted successfully!', 'success')
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/auction/go-live/<int:auction_id>', methods=['POST'])
@admin_required
def go_live_auction(auction_id):
    auction = Auction.query.get_or_404(auction_id)
    
    # Set auction to live state
    auction.status = 'live'
    auction.is_live = True
    auction.current_player_id = None
    auction.highest_bid = 0
    auction.highest_bid_team_id = None
    db.session.commit()
    
    # Broadcast auction started event to all connected users
    socketio.emit('auction_started', {
        'auction_id': auction.id,
        'auction_name': auction.name
    }, namespace='/')
    
    flash(f'Auction "{auction.name}" is now live!', 'success')
    return redirect(url_for('live_auction_control', auction_id=auction_id))

@app.route('/admin/auction/close/<int:auction_id>', methods=['POST'])
@admin_required
def close_auction(auction_id):
    auction = Auction.query.get_or_404(auction_id)
    auction.status = 'closed'
    auction.is_live = False
    db.session.commit()
    
    # Broadcast auction closed event
    socketio.emit('auction_closed', {
        'auction_id': auction.id,
        'auction_name': auction.name
    }, namespace='/')
    
    flash(f'Auction "{auction.name}" has been closed!', 'success')
    return redirect(url_for('admin_dashboard'))

# Live Auction Control Routes
@app.route('/admin/auction/live/<int:auction_id>')
@admin_required
def live_auction_control(auction_id):
    auction = Auction.query.get_or_404(auction_id)
    teams = Team.query.all()
    # Get available players (those in auction and not sold)
    # For now, all players in auction are available (we'll add sold_to field later if needed)
    available_players = auction.players if auction.players else []
    current_player = Player.query.get(auction.current_player_id) if auction.current_player_id else None
    
    return render_template('live_auction_control.html', 
                         auction=auction, 
                         teams=teams, 
                         available_players=available_players,
                         current_player=current_player)

@app.route('/admin/auction/select-player', methods=['POST'])
@admin_required
def select_player():
    try:
        if not request.is_json:
            return jsonify({'error': 'Request must be JSON'}), 400
        
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400
            
        auction_id = data.get('auction_id')
        player_id = data.get('player_id')
        is_random = data.get('random', False)
        
        if not auction_id:
            return jsonify({'error': 'Auction ID is required'}), 400
        
        auction = Auction.query.get(auction_id)
        if not auction:
            return jsonify({'error': 'Auction not found'}), 404
        
        if is_random:
            # Select random available player
            available_players = list(auction.players) if auction.players else []
            if not available_players:
                return jsonify({'error': 'No available players in this auction'}), 400
            player = random.choice(available_players)
            player_id = player.id
        else:
            if not player_id:
                return jsonify({'error': 'Player ID is required'}), 400
            player = Player.query.get(player_id)
            if not player:
                return jsonify({'error': 'Player not found'}), 404
            if player not in auction.players:
                return jsonify({'error': 'Player not in this auction'}), 400
        
        # Set current player
        auction.current_player_id = player_id
        auction.highest_bid = 0  # Start with 0, bid amount will be added to base
        auction.highest_bid_team_id = None
        db.session.commit()
        
        base_price = float(player.base_price) if player.base_price else 0.0
        
        # Broadcast player live event
        socketio.emit('player_live', {
            'player_id': player.id,
            'player_name': player.name,
            'base_price': base_price,
            'position': player.position or '',
            'batch': player.batch or '',
            'photo': player.photo_filename or '',
            'min_bid': float(auction.min_bid),
            'current_bid': 0.0,  # Start with 0
            'total_amount': base_price  # Total = base + bid (bid is 0 initially)
        }, namespace='/')
        
        return jsonify({
            'success': True,
            'player_id': player.id,
            'player_name': player.name
        })
    except Exception as e:
        db.session.rollback()
        print(f"Error in select_player: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Server error: {str(e)}'}), 500

@app.route('/admin/auction/place-bid', methods=['POST'])
@admin_required
def place_bid():
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'Invalid request data'}), 400
            
        auction_id = data.get('auction_id')
        player_id = data.get('player_id')
        team_id = data.get('team_id')
        bid_amount = data.get('bid_amount')
        
        if not all([auction_id, player_id, team_id, bid_amount]):
            return jsonify({'error': 'Missing required fields'}), 400
        
        try:
            bid_amount = float(bid_amount)
        except (ValueError, TypeError):
            return jsonify({'error': 'Invalid bid amount'}), 400
        
        auction = Auction.query.get_or_404(auction_id)
        team = Team.query.get_or_404(team_id)
        player = Player.query.get_or_404(player_id)
        
        # Basic validation only
        if not auction.is_live or auction.status != 'live':
            return jsonify({'error': 'Auction is not live'}), 400
        
        if auction.current_player_id != player_id:
            return jsonify({'error': 'This player is not currently being auctioned'}), 400
        
        # No amount validation - admin can put any amount
        # Create bid record
        new_bid = Bid(
            auction_id=auction_id,
            player_id=player_id,
            team_id=team_id,
            bid_amount=bid_amount
        )
        db.session.add(new_bid)
        
        # Update auction state - CUMULATIVE BIDDING: each bid ADDS to the total
        # highest_bid now stores the cumulative bid amount (not just the latest bid)
        auction.highest_bid = (auction.highest_bid or 0) + float(bid_amount)
        auction.highest_bid_team_id = team_id  # Last team to bid
        db.session.commit()
        
        # Calculate total (base_price + cumulative_bid_amount)
        total_amount = float(player.base_price) + float(auction.highest_bid)
        
        # Broadcast new bid event
        socketio.emit('new_bid', {
            'team_id': team_id,
            'team_name': team.name,
            'amount': bid_amount,  # This bid amount
            'cumulative_bid': float(auction.highest_bid),  # Total cumulative bids
            'base_price': float(player.base_price) if player.base_price else 0.0,
            'total_amount': total_amount,  # Base + Cumulative bids
            'player_id': player_id,
            'player_name': player.name
        }, namespace='/')
        
        return jsonify({
            'success': True,
            'bid_amount': bid_amount,
            'team_name': team.name
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Error placing bid: {str(e)}'}), 500

@app.route('/admin/auction/sell-player', methods=['POST'])
@admin_required
def sell_player():
    data = request.get_json()
    auction_id = data.get('auction_id')
    player_id = data.get('player_id')
    
    auction = Auction.query.get_or_404(auction_id)
    player = Player.query.get_or_404(player_id)
    
    if not auction.highest_bid_team_id:
        return jsonify({'error': 'No bid placed yet'}), 400
    
    # Store values before clearing
    sold_team_id = auction.highest_bid_team_id
    sold_team = Team.query.get(sold_team_id)
    cumulative_bid_amount = auction.highest_bid or 0
    base_price = float(player.base_price) if player.base_price else 0.0
    final_price = base_price + cumulative_bid_amount  # Final price = Base + All cumulative bids
    
    # Mark player as sold (we'll need to add these fields to Player model)
    # For now, we'll just update the auction state
    auction.current_player_id = None
    auction.highest_bid = 0
    auction.highest_bid_team_id = None
    db.session.commit()
    
    # Broadcast player sold event with final price
    socketio.emit('player_sold', {
        'player_id': player.id,
        'player_name': player.name,
        'team_id': sold_team_id,
        'team_name': sold_team.name if sold_team else None,
        'sold_price': final_price,  # Final price = Base + Cumulative bids
        'base_price': base_price,
        'cumulative_bids': cumulative_bid_amount
    }, namespace='/')
    
    return jsonify({'success': True})

@app.route('/admin/auction/team-spending/<int:auction_id>/<int:player_id>')
@admin_required
def get_team_spending(auction_id, player_id):
    """Get total spending by each team for a specific player"""
    try:
        from sqlalchemy import func
        
        # Get all bids for this player in this auction, grouped by team
        bids = db.session.query(
            Bid.team_id,
            Team.name,
            func.sum(Bid.bid_amount).label('total_bid')
        ).join(Team).filter(
            Bid.auction_id == auction_id,
            Bid.player_id == player_id
        ).group_by(Bid.team_id, Team.name).all()
        
        spending = []
        for bid in bids:
            spending.append({
                'team_id': bid.team_id,
                'team_name': bid.name,
                'total_bid': float(bid.total_bid)
            })
        
        return jsonify({
            'success': True,
            'spending': spending
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/auction/team-spending/<int:auction_id>/<int:player_id>')
def get_team_spending_public(auction_id, player_id):
    """Get total spending by each team for a specific player (public route)"""
    try:
        from sqlalchemy import func
        
        # Get all bids for this player in this auction, grouped by team
        bids = db.session.query(
            Bid.team_id,
            Team.name,
            func.sum(Bid.bid_amount).label('total_bid')
        ).join(Team).filter(
            Bid.auction_id == auction_id,
            Bid.player_id == player_id
        ).group_by(Bid.team_id, Team.name).all()
        
        spending = []
        for bid in bids:
            spending.append({
                'team_id': bid.team_id,
                'team_name': bid.name,
                'total_bid': float(bid.total_bid)
            })
        
        return jsonify({
            'success': True,
            'spending': spending
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Live Auction View for Teams
@app.route('/auction/live')
def live_auction_view():
    # Find active live auction
    auction = Auction.query.filter_by(is_live=True, status='live').first()
    if not auction:
        flash('No live auction currently', 'info')
        return redirect(url_for('home'))
    
    current_player = Player.query.get(auction.current_player_id) if auction.current_player_id else None
    highest_bid_team = Team.query.get(auction.highest_bid_team_id) if auction.highest_bid_team_id else None
    
    return render_template('live_auction_view.html', 
                         auction=auction,
                         current_player=current_player,
                         highest_bid_team=highest_bid_team)

if __name__ == '__main__':
    with app.app_context():
        # Create all tables first
        db.create_all()
        
        # Migrate existing database: Add new columns if they don't exist
        try:
            from sqlalchemy import inspect, text
            inspector = inspect(db.engine)
            
            # Check if team table exists
            if 'team' in inspector.get_table_names():
                columns = [col['name'] for col in inspector.get_columns('team')]
                
                # Add coowner_name column if it doesn't exist
                if 'coowner_name' not in columns:
                    db.session.execute(text('ALTER TABLE team ADD COLUMN coowner_name VARCHAR(100)'))
                    db.session.commit()
                    print("✓ Added coowner_name column to team table")
                
                # Add price column if it doesn't exist
                if 'price' not in columns:
                    db.session.execute(text('ALTER TABLE team ADD COLUMN price FLOAT DEFAULT 0.0'))
                    db.session.commit()
                    print("✓ Added price column to team table")
                
                # Add number_of_members column if it doesn't exist
                if 'number_of_members' not in columns:
                    db.session.execute(text('ALTER TABLE team ADD COLUMN number_of_members INTEGER DEFAULT 12'))
                    db.session.commit()
                    print("✓ Added number_of_members column to team table")
                
                # Add logo_filename column if it doesn't exist
                if 'logo_filename' not in columns:
                    db.session.execute(text('ALTER TABLE team ADD COLUMN logo_filename VARCHAR(255)'))
                    db.session.commit()
                    print("✓ Added logo_filename column to team table")
        except Exception as e:
            print(f"Migration check: {e}")
        
        # Migrate SlotManagement table
        try:
            if 'slot_management' in inspector.get_table_names():
                slot_columns = [col['name'] for col in inspector.get_columns('slot_management')]
                if 'total_teams' not in slot_columns:
                    db.session.execute(text('ALTER TABLE slot_management ADD COLUMN total_teams INTEGER DEFAULT 12'))
                    db.session.commit()
                    print("✓ Added total_teams column to slot_management table")
        except Exception as e:
            print(f"SlotManagement migration: {e}")
        
        # Migrate AuctionSetting table
        try:
            if 'auction_setting' in inspector.get_table_names():
                auction_columns = [col['name'] for col in inspector.get_columns('auction_setting')]
                if 'auction_date' not in auction_columns:
                    db.session.execute(text('ALTER TABLE auction_setting ADD COLUMN auction_date VARCHAR(100)'))
                    db.session.commit()
                    print("✓ Added auction_date column to auction_setting table")
                if 'auction_place' not in auction_columns:
                    db.session.execute(text('ALTER TABLE auction_setting ADD COLUMN auction_place VARCHAR(200)'))
                    db.session.commit()
                    print("✓ Added auction_place column to auction_setting table")
        except Exception as e:
            print(f"AuctionSetting migration: {e}")
        
        # Create Auction table if it doesn't exist
        try:
            if 'auction' not in inspector.get_table_names():
                print("✓ Creating auction table...")
            else:
                auction_columns = [col['name'] for col in inspector.get_columns('auction')]
                if 'auction_setting_id' not in auction_columns:
                    db.session.execute(text('ALTER TABLE auction ADD COLUMN auction_setting_id INTEGER'))
                    db.session.commit()
                    print("✓ Added auction_setting_id column to auction table")
                if 'is_live' not in auction_columns:
                    # SQLite uses INTEGER for boolean (0 or 1)
                    db.session.execute(text('ALTER TABLE auction ADD COLUMN is_live INTEGER DEFAULT 0'))
                    db.session.commit()
                    print("✓ Added is_live column to auction table")
                if 'current_player_id' not in auction_columns:
                    db.session.execute(text('ALTER TABLE auction ADD COLUMN current_player_id INTEGER'))
                    db.session.commit()
                    print("✓ Added current_player_id column to auction table")
                if 'highest_bid' not in auction_columns:
                    db.session.execute(text('ALTER TABLE auction ADD COLUMN highest_bid FLOAT DEFAULT 0.0'))
                    db.session.commit()
                    print("✓ Added highest_bid column to auction table")
                if 'highest_bid_team_id' not in auction_columns:
                    db.session.execute(text('ALTER TABLE auction ADD COLUMN highest_bid_team_id INTEGER'))
                    db.session.commit()
                    print("✓ Added highest_bid_team_id column to auction table")
        except Exception as e:
            print(f"Auction migration: {e}")
        
        # Create Bid table if it doesn't exist
        try:
            if 'bid' not in inspector.get_table_names():
                print("✓ Creating bid table...")
        except Exception as e:
            print(f"Bid migration: {e}")
        
        # Create player_photos directory
        os.makedirs(app.config['PLAYER_PHOTO_FOLDER'], exist_ok=True)
        
        # Create temp_photos directory for bulk uploads
        os.makedirs('static/temp_photos', exist_ok=True)
        
        # Initialize default admin if not exists
        if not Admin.query.first():
            admin = Admin(username='admin', password='admin123')
            db.session.add(admin)
        
        # Initialize slot management
        if not SlotManagement.query.first():
            slot_mgmt = SlotManagement(total_slots=12, total_teams=12, filled_slots=0, remaining_slots=12)
            db.session.add(slot_mgmt)
        
        # Initialize default auction time - Saturday Dec 14, 2025 at 10:00 AM
        if not AuctionSetting.query.first():
            default_time = datetime(2025, 12, 14, 10, 0, 0)
            auction_setting = AuctionSetting(
                auction_start_time=default_time,
                auction_date="Saturday, December 14, 2025",
                auction_place="Main Auditorium"
            )
            db.session.add(auction_setting)
        
        # Add sample teams if database is empty
        if not Team.query.first():
            sample_teams = [
                Team(name='Team Alpha', owner='John Doe', coowner_name='Jane Doe', batch='CSE 1', price=50000.00, number_of_members=12),
                Team(name='Team Beta', owner='Jane Smith', batch='CSE 2', price=45000.00, number_of_members=12),
                Team(name='Team Gamma', owner='Mike Johnson', coowner_name='Lisa Johnson', batch='CSE 1', price=55000.00, number_of_members=12),
                Team(name='Team Delta', owner='Sarah Williams', batch='CSE 3', price=48000.00, number_of_members=12),
            ]
            for team in sample_teams:
                db.session.add(team)
            
            # Add sample team users
            team1 = Team.query.filter_by(name='Team Alpha').first()
            if team1:
                team_user = TeamUser(team_id=team1.id, username='team1', password='team123')
                db.session.add(team_user)
        
        db.session.commit()
    
    # Only run with Flask dev server if not in production
    if os.environ.get('FLASK_ENV') != 'production':
        socketio.run(app, debug=True, host='0.0.0.0', port=5000, allow_unsafe_werkzeug=True, use_reloader=False)

